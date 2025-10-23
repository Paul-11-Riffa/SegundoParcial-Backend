import django_filters
from rest_framework import views, response, status, permissions, generics
from .models import Order, OrderItem, Product
from .serializers import OrderSerializer
from django.conf import settings
import stripe
from django.http import HttpResponse
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import Table, TableStyle
from reportlab.lib.units import inch
from .filters import OrderFilter
from datetime import datetime, timedelta
from django.utils import timezone
import re
class CartView(views.APIView):
    """
    Vista para gestionar el carrito de compras del usuario.
    """
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        """
        Obtiene o crea el carrito de compras actual (en estado 'PENDING') del usuario.
        """
        # Asegurémonos de obtener solo carritos sin items completados
        cart = Order.objects.filter(customer=request.user, status='PENDING').first()
        if not cart:
            cart = Order.objects.create(customer=request.user, status='PENDING', total_price=0.00)
        serializer = OrderSerializer(cart)
        return response.Response(serializer.data)

    def post(self, request):
        """
        Añade un producto al carrito.
        Espera recibir: { "product_id": <id>, "quantity": <cantidad> }
        """
        product_id = request.data.get('product_id')
        quantity = int(request.data.get('quantity', 1))

        if not product_id:
            return response.Response({'error': 'Product ID is required.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            product = Product.objects.get(id=product_id)
        except Product.DoesNotExist:
            return response.Response({'error': 'Product not found.'}, status=status.HTTP_404_NOT_FOUND)

        cart, _ = Order.objects.get_or_create(customer=request.user, status='PENDING')

        # Si el producto ya está en el carrito, actualiza la cantidad. Si no, lo crea.
        order_item, created = OrderItem.objects.get_or_create(order=cart, product=product, defaults={'price': product.price})

        if not created:
            order_item.quantity += quantity
        else:
            order_item.quantity = quantity

        # Valida que haya suficiente stock
        if order_item.quantity > product.stock:
            return response.Response({'error': 'Not enough stock available.'}, status=status.HTTP_400_BAD_REQUEST)

        order_item.save()

        # Recalcula el precio total del carrito
        cart.total_price = sum(item.price * item.quantity for item in cart.items.all())
        cart.save()

        serializer = OrderSerializer(cart)
        return response.Response(serializer.data, status=status.HTTP_200_OK)


class CartItemView(views.APIView):
    """
    Vista para actualizar o eliminar un artículo específico del carrito.
    """
    permission_classes = [permissions.IsAuthenticated]

    def put(self, request, item_id):
        """
        Actualiza la cantidad de un artículo en el carrito.
        Espera recibir: { "quantity": <nueva_cantidad> }
        """
        quantity = int(request.data.get('quantity', 1))

        try:
            order_item = OrderItem.objects.get(id=item_id, order__customer=request.user)
        except OrderItem.DoesNotExist:
            return response.Response({'error': 'Cart item not found.'}, status=status.HTTP_404_NOT_FOUND)

        if quantity > 0:
            if quantity > order_item.product.stock:
                return response.Response({'error': 'Not enough stock available.'}, status=status.HTTP_400_BAD_REQUEST)
            order_item.quantity = quantity
            order_item.save()
        else: # Si la cantidad es 0, elimina el artículo
            order_item.delete()

        # Recalcula el precio total
        cart = order_item.order
        cart.total_price = sum(item.price * item.quantity for item in cart.items.all())
        cart.save()

        serializer = OrderSerializer(cart)
        return response.Response(serializer.data)

    def delete(self, request, item_id):
        """
        Elimina un artículo del carrito.
        """
        try:
            order_item = OrderItem.objects.get(id=item_id, order__customer=request.user)
        except OrderItem.DoesNotExist:
            return response.Response({'error': 'Cart item not found.'}, status=status.HTTP_404_NOT_FOUND)

        cart = order_item.order
        order_item.delete()

        # Recalcula el precio total
        cart.total_price = sum(item.price * item.quantity for item in cart.items.all())
        cart.save()

        serializer = OrderSerializer(cart)
        return response.Response(serializer.data)

class StripeCheckoutView(views.APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        """
        Crea una sesión de pago en Stripe con los artículos del carrito.
        """
        # Asigna la clave secreta de Stripe desde la configuración
        stripe.api_key = settings.STRIPE_SECRET_KEY

        try:
            # 1. Obtiene el carrito del usuario
            cart = Order.objects.get(customer=request.user, status='PENDING')
            if not cart.items.exists():
                return response.Response({'error': 'Your cart is empty.'}, status=status.HTTP_400_BAD_REQUEST)

            # 2. Prepara la lista de productos para Stripe
            line_items = []
            for item in cart.items.all():
                line_items.append({
                    'price_data': {
                        'currency': 'usd', # Puedes cambiarlo a tu moneda local (ej. 'bob')
                        'product_data': {
                            'name': item.product.name,
                        },
                        'unit_amount': int(item.product.price * 100), # Stripe necesita el precio en centavos
                    },
                    'quantity': item.quantity,
                })

            # 3. Define las URLs de éxito y cancelación
            # (Estas son las páginas a las que Stripe redirigirá al usuario después del pago)
            frontend_base_url = "http://localhost:3000"
            success_url = f"{frontend_base_url}/order/success?session_id={{CHECKOUT_SESSION_ID}}"  # Pasamos el ID de sesión para verificación opcional
            cancel_url = f"{frontend_base_url}/order/cancel"

            # 4. Crea la sesión de checkout en Stripe
            checkout_session = stripe.checkout.Session.create(
                payment_method_types=['card'],
                line_items=line_items,
                mode='payment',
                success_url=success_url,
                cancel_url=cancel_url,
                # Guarda el ID de nuestra orden en los metadatos de Stripe
                # ¡Esto es crucial para saber qué orden se pagó!
                metadata={
                    'order_id': cart.id,
                    'user_id': request.user.id
                }
            )

            # 5. IMPORTANTE: Cambia el estado del carrito a PROCESSING inmediatamente
            # Esto evita que el usuario vea el carrito mientras el pago está en proceso
            cart.status = 'PROCESSING'
            cart.save()

            # 6. Devuelve la URL de la sesión de pago al frontend
            return response.Response({'checkout_url': checkout_session.url})

        except Order.DoesNotExist:
            return response.Response({'error': 'You do not have an active cart.'}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return response.Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class StripeWebhookView(views.APIView):
    """
    Escucha los eventos de Stripe, específicamente cuando un pago es exitoso.
    """
    permission_classes = [permissions.AllowAny] # Debe ser accesible públicamente para Stripe

    def post(self, request):
        payload = request.body
        sig_header = request.META.get('HTTP_STRIPE_SIGNATURE')
        endpoint_secret = settings.STRIPE_WEBHOOK_SECRET
        event = None

        try:
            # 1. Verifica que el evento realmente venga de Stripe
            event = stripe.Webhook.construct_event(
                payload, sig_header, endpoint_secret
            )
        except ValueError as e:
            # Payload inválido
            return response.Response(status=status.HTTP_400_BAD_REQUEST)
        except stripe.error.SignatureVerificationError as e:
            # Firma inválida
            return response.Response(status=status.HTTP_400_BAD_REQUEST)

        # 2. Maneja el evento específico de "pago completado"
        if event['type'] == 'checkout.session.completed':
            session = event['data']['object']

            # Obtiene el ID de nuestra orden que guardamos en los metadatos
            order_id = session.get('metadata', {}).get('order_id')

            if order_id is None:
                return response.Response({'error': 'Missing order_id in Stripe metadata'}, status=status.HTTP_400_BAD_REQUEST)

            try:
                # 3. Encuentra la orden y actualiza su estado
                order = Order.objects.get(id=order_id, status='PROCESSING')
                order.status = 'COMPLETED'
                order.save()

                # 4. Reduce el stock de los productos vendidos
                for item in order.items.all():
                    product = item.product
                    if product.stock >= item.quantity:
                        product.stock -= item.quantity
                        product.save()
                    else:
                        # Manejar el caso de que no haya suficiente stock (raro, pero posible)
                        print(f"Alerta: Stock insuficiente para el producto {product.id} en la orden {order.id}")
                        # Aquí podrías enviar un email de alerta al administrador

            except Order.DoesNotExist:
                return response.Response({'error': f'Order with ID {order_id} not found or already processed.'}, status=status.HTTP_404_NOT_FOUND)

        # Si es otro tipo de evento, simplemente lo ignoramos por ahora
        else:
            print(f"Evento no manejado: {event['type']}")

        # 5. Responde a Stripe para confirmar que recibimos el evento
        return response.Response(status=status.HTTP_200_OK)

# --- VISTA PARA COMPLETAR LA ORDEN DEL USUARIO AUTENTICADO ---
class CompleteOrderView(views.APIView):
    """
    Endpoint para completar la orden pendiente del usuario autenticado.
    Se llama desde el frontend después de un pago exitoso en Stripe.
    """
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        try:
            # 1. Buscar orden en PROCESSING (después del checkout en Stripe)
            cart = Order.objects.filter(customer=request.user, status='PROCESSING').first()
            
            if not cart:
                # Si no hay orden en PROCESSING, buscar en PENDING (por compatibilidad)
                cart = Order.objects.filter(customer=request.user, status='PENDING').first()
            
            if not cart:
                return response.Response({
                    'error': 'No order found to complete'
                }, status=status.HTTP_404_NOT_FOUND)

            # 2. Cambia el estado a COMPLETADO
            cart.status = 'COMPLETED'
            cart.save()

            # 3. Reduce el stock de los productos
            for item in cart.items.all():
                product = item.product
                if product.stock >= item.quantity:
                    product.stock -= item.quantity
                    product.save()
                else:
                    print(f"Alerta de Stock: Stock insuficiente para el producto {product.id}")

            # 4. ✅ CREAR NUEVO CARRITO VACÍO PARA EL USUARIO
            Order.objects.create(customer=request.user, status='PENDING', total_price=0.00)

            return response.Response({
                'success': True,
                'message': 'Order completed successfully',
                'order_id': cart.id
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return response.Response({
                'error': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

# --- VISTA DE PRUEBA PARA SIMULAR EL WEBHOOK MANUALMENTE ---
class ManualOrderCompletionView(views.APIView):
    """
    [SOLO PARA DESARROLLO]
    Endpoint para forzar la finalización de la orden pendiente de un usuario.
    Simula el comportamiento del webhook de Stripe.
    """
    permission_classes = [permissions.IsAdminUser] # Protegido para que solo un admin pueda usarlo

    def post(self, request):
        user_id = request.data.get('user_id')
        if not user_id:
            return response.Response({'error': 'user_id is required.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            # 1. Encuentra la orden PENDIENTE del cliente especificado
            order = Order.objects.get(customer_id=user_id, status='PENDING')

            # 2. Cambia el estado a COMPLETADO
            order.status = 'COMPLETED'
            order.save()

            # 3. Reduce el stock de los productos
            for item in order.items.all():
                product = item.product
                if product.stock >= item.quantity:
                    product.stock -= item.quantity
                    product.save()
                else:
                    print(f"Alerta de Stock (Debug): Stock insuficiente para el producto {product.id}")

            return response.Response({'success': f'Order {order.id} for user {user_id} has been marked as COMPLETED.'}, status=status.HTTP_200_OK)

        except Order.DoesNotExist:
            return response.Response({'error': f'No pending order found for user {user_id}.'}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return response.Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class CompleteUserOrderView(views.APIView):
    """
    [DESARROLLO] Endpoint para completar la orden en PROCESSING del usuario autenticado.
    Se llama cuando el usuario regresa de Stripe exitosamente.
    """
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        try:
            # Busca la orden en estado PROCESSING del usuario
            order = Order.objects.get(customer=request.user, status='PROCESSING')
            
            # Cambia el estado a COMPLETED
            order.status = 'COMPLETED'
            order.save()
            
            # Reduce el stock de los productos
            for item in order.items.all():
                product = item.product
                if product.stock >= item.quantity:
                    product.stock -= item.quantity
                    product.save()
                else:
                    print(f"Alerta: Stock insuficiente para el producto {product.id} en la orden {order.id}")
            
            return response.Response({
                'success': True,
                'message': 'Orden completada exitosamente',
                'order_id': order.id
            }, status=status.HTTP_200_OK)
            
        except Order.DoesNotExist:
            return response.Response({
                'success': False,
                'message': 'No se encontró una orden en proceso'
            }, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return response.Response({
                'success': False,
                'error': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

# --- NUEVA VISTA PARA EL HISTORIAL DE VENTAS (SOLO ADMINS) ---
class SalesHistoryView(generics.ListAPIView):
    """
    Endpoint para que los administradores vean todas las órdenes completadas.
    """
    permission_classes = [permissions.IsAdminUser]
    serializer_class = OrderSerializer
    filterset_class = OrderFilter
    filter_backends = [django_filters.rest_framework.DjangoFilterBackend]

    def get_queryset(self):
        """
        Filtra las órdenes para devolver solo las que tienen el estado 'COMPLETED'.
        """
        return Order.objects.filter(status='COMPLETED').order_by('-updated_at')

# --- VISTA PARA GENERAR COMPROBANTES EN PDF (ADMIN) ---
class GenerateOrderReceiptPDF(views.APIView):
    permission_classes = [permissions.IsAdminUser]

    def get(self, request, order_id):
        try:
            # 1. Buscamos la orden completada
            order = Order.objects.get(id=order_id, status='COMPLETED')
        except Order.DoesNotExist:
            return response.Response({'error': 'Completed order not found.'}, status=status.HTTP_404_NOT_FOUND)

        # 2. Creamos una respuesta HTTP de tipo PDF
        response_pdf = HttpResponse(content_type='application/pdf')
        response_pdf['Content-Disposition'] = f'attachment; filename="comprobante_venta_{order.id}.pdf"'

        # 3. Creamos el lienzo del PDF
        p = canvas.Canvas(response_pdf, pagesize=letter)
        width, height = letter

        # --- ENCABEZADO CON LOGO Y EMPRESA ---
        # Nombre de la empresa grande y destacado
        p.setFont("Helvetica-Bold", 32)
        p.setFillColor(colors.HexColor('#667eea'))
        p.drawCentredString(width / 2, height - 60, "SMARTSALES365")
        
        # Línea decorativa bajo el nombre
        p.setStrokeColor(colors.HexColor('#764ba2'))
        p.setLineWidth(3)
        p.line(72, height - 75, width - 72, height - 75)
        
        # Subtítulo de la empresa
        p.setFont("Helvetica", 11)
        p.setFillColor(colors.HexColor('#495057'))
        p.drawCentredString(width / 2, height - 95, "Sistema Inteligente de Gestión Comercial")
        
        # --- INFORMACIÓN DEL COMPROBANTE ---
        p.setFont("Helvetica-Bold", 18)
        p.setFillColor(colors.HexColor('#1a222e'))
        p.drawCentredString(width / 2, height - 130, "COMPROBANTE DE VENTA")
        
        # Recuadro con información de la orden
        p.setStrokeColor(colors.HexColor('#667eea'))
        p.setLineWidth(1.5)
        p.rect(72, height - 200, (width - 144) / 2 - 10, 50, stroke=1, fill=0)
        p.rect(width / 2 + 10, height - 200, (width - 144) / 2 - 10, 50, stroke=1, fill=0)
        
        p.setFont("Helvetica", 10)
        p.setFillColor(colors.HexColor('#667eea'))
        p.drawString(82, height - 160, "N° DE ORDEN:")
        p.drawString(width / 2 + 20, height - 160, "FECHA DE EMISIÓN:")
        
        p.setFont("Helvetica-Bold", 14)
        p.setFillColor(colors.black)
        p.drawString(82, height - 180, f"#{order.id:05d}")
        p.setFont("Helvetica-Bold", 12)
        p.drawString(width / 2 + 20, height - 180, order.updated_at.strftime('%d/%m/%Y %H:%M'))

        # --- INFORMACIÓN DEL CLIENTE ---
        p.setStrokeColor(colors.HexColor('#667eea'))
        p.rect(72, height - 260, width - 144, 45, stroke=1, fill=0)
        
        p.setFont("Helvetica", 10)
        p.setFillColor(colors.HexColor('#667eea'))
        p.drawString(82, height - 225, "CLIENTE:")
        
        customer_name = f"{order.customer.first_name} {order.customer.last_name}".strip()
        if not customer_name:
            customer_name = order.customer.username
        
        p.setFont("Helvetica-Bold", 13)
        p.setFillColor(colors.black)
        p.drawString(82, height - 243, customer_name)
        p.setFont("Helvetica", 10)
        p.setFillColor(colors.HexColor('#6c757d'))
        p.drawString(82, height - 255, f"Email: {order.customer.email}")

        # --- TABLA DE PRODUCTOS ---
        table_data = [['PRODUCTO', 'CANT.', 'PRECIO UNIT.', 'SUBTOTAL']]
        for item in order.items.all():
            subtotal = item.quantity * item.price
            table_data.append([
                item.product.name,
                str(item.quantity),
                f"${item.price:.2f}",
                f"${subtotal:.2f}"
            ])

        table = Table(table_data, colWidths=[3.5 * inch, 0.8 * inch, 1.2 * inch, 1.3 * inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('TOPPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 1.5, colors.HexColor('#e9ecef')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))

        table_height = table.wrap(width, height)[1]
        table.drawOn(p, 72, height - 290 - table_height)

        # --- TOTAL ---
        # Recuadro para el total
        total_y = height - 310 - table_height
        p.setFillColor(colors.HexColor('#667eea'))
        p.rect(width - 252, total_y - 45, 180, 40, stroke=0, fill=1)
        
        p.setFont("Helvetica-Bold", 12)
        p.setFillColor(colors.white)
        p.drawString(width - 240, total_y - 20, "TOTAL A PAGAR:")
        p.setFont("Helvetica-Bold", 18)
        p.drawString(width - 240, total_y - 38, f"${order.total_price:.2f} USD")

        # --- PIE DE PÁGINA ---
        footer_y = 100
        p.setFont("Helvetica", 9)
        p.setFillColor(colors.HexColor('#6c757d'))
        p.drawCentredString(width / 2, footer_y, "Gracias por su compra - SmartSales365")
        p.drawCentredString(width / 2, footer_y - 15, "www.smartsales365.com | contacto@smartsales365.com")
        
        # Línea decorativa en el pie
        p.setStrokeColor(colors.HexColor('#764ba2'))
        p.setLineWidth(2)
        p.line(72, footer_y - 30, width - 72, footer_y - 30)

        p.showPage()
        p.save()

        return response_pdf


# --- VISTA PARA GENERAR COMPROBANTES EN PDF (CLIENTE) ---
class GenerateMyOrderReceiptPDF(views.APIView):
    """
    Endpoint para que un cliente genere el comprobante de su propia orden.
    Solo puede acceder a sus propias órdenes completadas.
    """
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, order_id):
        try:
            # 1. Buscamos la orden completada que pertenezca al usuario actual
            order = Order.objects.get(id=order_id, status='COMPLETED', customer=request.user)
        except Order.DoesNotExist:
            return response.Response({'error': 'Orden completada no encontrada o no tienes acceso a ella.'}, status=status.HTTP_404_NOT_FOUND)

        # 2. Creamos una respuesta HTTP de tipo PDF
        response_pdf = HttpResponse(content_type='application/pdf')
        response_pdf['Content-Disposition'] = f'attachment; filename="comprobante_orden_{order.id}.pdf"'

        # 3. Creamos el lienzo del PDF
        p = canvas.Canvas(response_pdf, pagesize=letter)
        width, height = letter

        # --- ENCABEZADO CON LOGO Y EMPRESA ---
        # Nombre de la empresa grande y destacado
        p.setFont("Helvetica-Bold", 32)
        p.setFillColor(colors.HexColor('#667eea'))
        p.drawCentredString(width / 2, height - 60, "SMARTSALES365")
        
        # Línea decorativa bajo el nombre
        p.setStrokeColor(colors.HexColor('#764ba2'))
        p.setLineWidth(3)
        p.line(72, height - 75, width - 72, height - 75)
        
        # Subtítulo de la empresa
        p.setFont("Helvetica", 11)
        p.setFillColor(colors.HexColor('#495057'))
        p.drawCentredString(width / 2, height - 95, "Sistema Inteligente de Gestión Comercial")
        
        # --- INFORMACIÓN DEL COMPROBANTE ---
        p.setFont("Helvetica-Bold", 18)
        p.setFillColor(colors.HexColor('#1a222e'))
        p.drawCentredString(width / 2, height - 130, "COMPROBANTE DE COMPRA")
        
        # Recuadro con información de la orden
        p.setStrokeColor(colors.HexColor('#667eea'))
        p.setLineWidth(1.5)
        p.rect(72, height - 200, (width - 144) / 2 - 10, 50, stroke=1, fill=0)
        p.rect(width / 2 + 10, height - 200, (width - 144) / 2 - 10, 50, stroke=1, fill=0)
        
        p.setFont("Helvetica", 10)
        p.setFillColor(colors.HexColor('#667eea'))
        p.drawString(82, height - 160, "N° DE ORDEN:")
        p.drawString(width / 2 + 20, height - 160, "FECHA DE COMPRA:")
        
        p.setFont("Helvetica-Bold", 14)
        p.setFillColor(colors.black)
        p.drawString(82, height - 180, f"#{order.id:05d}")
        p.setFont("Helvetica-Bold", 12)
        p.drawString(width / 2 + 20, height - 180, order.updated_at.strftime('%d/%m/%Y %H:%M'))

        # --- INFORMACIÓN DEL CLIENTE ---
        p.setStrokeColor(colors.HexColor('#667eea'))
        p.rect(72, height - 260, width - 144, 45, stroke=1, fill=0)
        
        p.setFont("Helvetica", 10)
        p.setFillColor(colors.HexColor('#667eea'))
        p.drawString(82, height - 225, "DATOS DEL CLIENTE:")
        
        customer_name = f"{order.customer.first_name} {order.customer.last_name}".strip()
        if not customer_name:
            customer_name = order.customer.username
        
        p.setFont("Helvetica-Bold", 13)
        p.setFillColor(colors.black)
        p.drawString(82, height - 243, customer_name)
        p.setFont("Helvetica", 10)
        p.setFillColor(colors.HexColor('#6c757d'))
        p.drawString(82, height - 255, f"Email: {order.customer.email}")

        # --- TABLA DE PRODUCTOS ---
        table_data = [['PRODUCTO', 'CANT.', 'PRECIO UNIT.', 'SUBTOTAL']]
        for item in order.items.all():
            subtotal = item.quantity * item.price
            table_data.append([
                item.product.name,
                str(item.quantity),
                f"${item.price:.2f}",
                f"${subtotal:.2f}"
            ])

        table = Table(table_data, colWidths=[3.5 * inch, 0.8 * inch, 1.2 * inch, 1.3 * inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('TOPPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 1.5, colors.HexColor('#e9ecef')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))

        table_height = table.wrap(width, height)[1]
        table.drawOn(p, 72, height - 290 - table_height)

        # --- TOTAL ---
        # Recuadro para el total
        total_y = height - 310 - table_height
        p.setFillColor(colors.HexColor('#667eea'))
        p.rect(width - 252, total_y - 45, 180, 40, stroke=0, fill=1)
        
        p.setFont("Helvetica-Bold", 12)
        p.setFillColor(colors.white)
        p.drawString(width - 240, total_y - 20, "TOTAL PAGADO:")
        p.setFont("Helvetica-Bold", 18)
        p.drawString(width - 240, total_y - 38, f"${order.total_price:.2f} USD")

        # --- PIE DE PÁGINA ---
        footer_y = 100
        p.setFont("Helvetica", 9)
        p.setFillColor(colors.HexColor('#6c757d'))
        p.drawCentredString(width / 2, footer_y, "¡Gracias por su compra! - SmartSales365")
        p.drawCentredString(width / 2, footer_y - 15, "www.smartsales365.com | contacto@smartsales365.com")
        p.drawCentredString(width / 2, footer_y - 30, "Este documento es un comprobante válido de su compra")
        
        # Línea decorativa en el pie
        p.setStrokeColor(colors.HexColor('#764ba2'))
        p.setLineWidth(2)
        p.line(72, footer_y - 45, width - 72, footer_y - 45)

        p.showPage()
        p.save()

        return response_pdf

# --- NUEVA VISTA PARA LAS ÓRDENES DEL CLIENTE LOGUEADO ---
class MyOrderListView(generics.ListAPIView):
    """
    Endpoint para que un cliente vea su propio historial de órdenes
    (incluyendo carritos pendientes y ventas completadas).
    """
    permission_classes = [permissions.IsAuthenticated] # Solo usuarios logueados
    serializer_class = OrderSerializer

    def get_queryset(self):
        """
        Filtra las órdenes para devolver solo las del usuario actual,
        excluyendo carritos pendientes Y vacíos.
        """
        user = self.request.user
        # Obtenemos todas las órdenes del usuario que NO sean PENDING
        # (solo mostramos las COMPLETED y CANCELLED, no el carrito activo)
        queryset = Order.objects.filter(customer=user).exclude(status='PENDING')
        # Ordenamos por fecha, las más recientes primero
        return queryset.order_by('-created_at')


# --- VISTA PARA GENERAR REPORTES DINÁMICOS ---
class GenerateDynamicReportView(views.APIView):
    """
    Vista para generar reportes dinámicos basados en comandos de texto o voz.
    Interpreta el prompt del usuario y genera el reporte solicitado en PDF, Excel o pantalla.
    """
    permission_classes = [permissions.IsAuthenticated]
    
    def post(self, request):
        prompt = request.data.get('prompt', '').lower()
        
        if not prompt:
            return response.Response(
                {'detail': 'Se requiere un prompt para generar el reporte'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            # Parsear el prompt para extraer información
            parsed_data = self._parse_prompt(prompt)
            
            # Obtener las órdenes filtradas
            orders = self._get_filtered_orders(parsed_data)
            
            # Determinar el formato de salida
            output_format = parsed_data.get('format', 'screen')
            
            if output_format == 'pdf':
                return self._generate_pdf_report(orders, parsed_data)
            elif output_format == 'excel':
                return self._generate_excel_report(orders, parsed_data)
            else:
                return self._generate_screen_report(orders, parsed_data)
                
        except Exception as e:
            print(f"⚠️ {str(e)}")
            return response.Response(
                {'detail': f'Error al generar el reporte: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    def _parse_prompt(self, prompt):
        """
        Parsea el prompt para extraer fechas, formato y otros parámetros.
        """
        parsed = {
            'format': 'screen',
            'start_date': None,
            'end_date': None,
            'group_by': None
        }
        
        # Detectar formato
        if 'pdf' in prompt:
            parsed['format'] = 'pdf'
        elif 'excel' in prompt:
            parsed['format'] = 'excel'
        elif 'pantalla' in prompt or 'screen' in prompt:
            parsed['format'] = 'screen'
        
        # Detectar agrupación
        if 'producto' in prompt or 'product' in prompt:
            parsed['group_by'] = 'product'
        elif 'cliente' in prompt or 'customer' in prompt:
            parsed['group_by'] = 'customer'
        elif 'categoría' in prompt or 'category' in prompt:
            parsed['group_by'] = 'category'
        
        # Detectar fechas - Formato DD/MM/YYYY
        date_pattern = r'(\d{1,2})/(\d{1,2})/(\d{4})'
        dates = re.findall(date_pattern, prompt)
        
        if len(dates) >= 2:
            # Hay fecha inicio y fin
            try:
                start = datetime(int(dates[0][2]), int(dates[0][1]), int(dates[0][0]))
                end = datetime(int(dates[1][2]), int(dates[1][1]), int(dates[1][0]), 23, 59, 59)
                parsed['start_date'] = timezone.make_aware(start)
                parsed['end_date'] = timezone.make_aware(end)
            except:
                pass
        elif len(dates) == 1:
            # Solo una fecha
            try:
                date = datetime(int(dates[0][2]), int(dates[0][1]), int(dates[0][0]))
                parsed['start_date'] = timezone.make_aware(date.replace(hour=0, minute=0, second=0))
                parsed['end_date'] = timezone.make_aware(date.replace(hour=23, minute=59, second=59))
            except:
                pass
        
        # Detectar meses
        months = {
            'enero': 1, 'febrero': 2, 'marzo': 3, 'abril': 4,
            'mayo': 5, 'junio': 6, 'julio': 7, 'agosto': 8,
            'septiembre': 9, 'octubre': 10, 'noviembre': 11, 'diciembre': 12
        }
        
        for month_name, month_num in months.items():
            if month_name in prompt:
                # Detectar año
                year_match = re.search(r'\b(202\d)\b', prompt)
                year = int(year_match.group(1)) if year_match else timezone.now().year
                
                # Primer día del mes
                start = datetime(year, month_num, 1)
                # Último día del mes
                if month_num == 12:
                    end = datetime(year, 12, 31, 23, 59, 59)
                else:
                    end = datetime(year, month_num + 1, 1) - timedelta(seconds=1)
                
                parsed['start_date'] = timezone.make_aware(start)
                parsed['end_date'] = timezone.make_aware(end)
                break
        
        # Detectar "año"
        if 'año' in prompt or 'year' in prompt:
            year_match = re.search(r'\b(202\d)\b', prompt)
            if year_match:
                year = int(year_match.group(1))
                parsed['start_date'] = timezone.make_aware(datetime(year, 1, 1))
                parsed['end_date'] = timezone.make_aware(datetime(year, 12, 31, 23, 59, 59))
        
        return parsed
    
    def _get_filtered_orders(self, parsed_data):
        """
        Obtiene las órdenes filtradas según los parámetros parseados.
        """
        orders = Order.objects.filter(status='COMPLETED')
        
        if parsed_data['start_date']:
            orders = orders.filter(updated_at__gte=parsed_data['start_date'])
        
        if parsed_data['end_date']:
            orders = orders.filter(updated_at__lte=parsed_data['end_date'])
        
        return orders.order_by('-updated_at')
    
    def _generate_pdf_report(self, orders, parsed_data):
        """
        Genera un reporte en formato PDF.
        """
        from reportlab.pdfgen import canvas
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.lib import colors
        from reportlab.platypus import Table, TableStyle
        from reportlab.lib.units import inch
        
        response_pdf = HttpResponse(content_type='application/pdf')
        response_pdf['Content-Disposition'] = 'attachment; filename="reporte_ventas.pdf"'
        
        p = canvas.Canvas(response_pdf, pagesize=letter)
        width, height = letter
        
        # Título
        p.setFont("Helvetica-Bold", 18)
        p.drawString(72, height - 50, "Reporte de Ventas")
        
        # Fechas
        p.setFont("Helvetica", 12)
        y_position = height - 80
        
        if parsed_data['start_date'] and parsed_data['end_date']:
            p.drawString(72, y_position, 
                f"Período: {parsed_data['start_date'].strftime('%d/%m/%Y')} - {parsed_data['end_date'].strftime('%d/%m/%Y')}")
            y_position -= 25
        
        # Estadísticas generales
        total_ventas = sum(order.total_price for order in orders)
        p.drawString(72, y_position, f"Total de órdenes: {orders.count()}")
        y_position -= 20
        p.drawString(72, y_position, f"Monto total: ${total_ventas:.2f}")
        y_position -= 30
        
        # Tabla de órdenes
        table_data = [['Orden #', 'Cliente', 'Fecha', 'Total']]
        
        for order in orders[:15]:  # Limitar a 15 para que quepa en una página
            table_data.append([
                str(order.id),
                order.customer.username,
                order.updated_at.strftime('%d/%m/%Y'),
                f"${order.total_price:.2f}"
            ])
        
        if table_data:
            table = Table(table_data, colWidths=[1*inch, 2*inch, 1.5*inch, 1.5*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1A222E')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 10),
            ]))
            
            table_height = table.wrap(width, height)[1]
            table.drawOn(p, 72, y_position - table_height - 20)
        
        p.showPage()
        p.save()
        
        return response_pdf
    
    def _generate_excel_report(self, orders, parsed_data):
        """
        Genera un reporte en formato Excel.
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Reporte de Ventas"
            
            # Encabezados
            headers = ['Orden #', 'Cliente', 'Email', 'Fecha', 'Total', 'Items']
            ws.append(headers)
            
            # Estilo de encabezados
            header_fill = PatternFill(start_color="1A222E", end_color="1A222E", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
            
            # Datos
            for order in orders:
                items_count = order.items.count()
                ws.append([
                    order.id,
                    order.customer.username,
                    order.customer.email,
                    order.updated_at.strftime('%d/%m/%Y %H:%M'),
                    float(order.total_price),
                    items_count
                ])
            
            # Ajustar ancho de columnas
            ws.column_dimensions['A'].width = 10
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 30
            ws.column_dimensions['D'].width = 18
            ws.column_dimensions['E'].width = 12
            ws.column_dimensions['F'].width = 10
            
            # Guardar en memoria
            from io import BytesIO
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            response_excel = HttpResponse(
                buffer.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response_excel['Content-Disposition'] = 'attachment; filename="reporte_ventas.xlsx"'
            
            return response_excel
            
        except ImportError:
            return response.Response(
                {'detail': 'El módulo openpyxl no está instalado. Por favor instálalo con: pip install openpyxl'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    def _generate_screen_report(self, orders, parsed_data):
        """
        Genera un reporte para mostrar en pantalla (JSON).
        """
        data = {
            'total_orders': orders.count(),
            'total_amount': sum(order.total_price for order in orders),
            'period': {
                'start': parsed_data['start_date'].strftime('%d/%m/%Y') if parsed_data['start_date'] else None,
                'end': parsed_data['end_date'].strftime('%d/%m/%Y') if parsed_data['end_date'] else None,
            },
            'orders': []
        }
        
        for order in orders[:50]:  # Limitar a 50 órdenes
            order_data = {
                'id': order.id,
                'customer': order.customer.username,
                'email': order.customer.email,
                'date': order.updated_at.strftime('%d/%m/%Y %H:%M'),
                'total': float(order.total_price),
                'items': [
                    {
                        'product': item.product.name,
                        'quantity': item.quantity,
                        'price': float(item.price)
                    }
                    for item in order.items.all()
                ]
            }
            data['orders'].append(order_data)
        
        return response.Response(data, status=status.HTTP_200_OK)
